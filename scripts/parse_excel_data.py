#!/usr/bin/env python3
"""
Parse MVHS Football Excel/CSV files into structured JSON data.

Reads:
  1. All County Winners.xlsx
  2. All League Winners.xlsx
  3. All State Recognition.xlsx
  4. Cal Hi All State Players.xlsx
  5. CIF Cal Hi Records.xlsx
  6. Copy of Roster with sizes.xlsx
  7. Mission Viejo Football (2026 - 2027) Contact Info.csv

Outputs: parsed_data.json with sections: honors, records, roster, contacts
"""

import json
import csv
import os
import re
import datetime
import openpyxl

class SafeEncoder(json.JSONEncoder):
    """JSON encoder that handles datetime objects."""
    def default(self, obj):
        if isinstance(obj, (datetime.datetime, datetime.date)):
            return obj.isoformat()
        if isinstance(obj, datetime.time):
            return obj.isoformat()
        return super().default(obj)


DOWNLOADS = "/Users/markcoleman/Downloads"
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "parsed_data.json")


def is_empty(value):
    """Check if a cell value should be treated as empty/no-data."""
    if value is None:
        return True
    s = str(value).strip()
    if s == "" or s == "None" or s == "none" or s == "N/A":
        return True
    return False


def clean_text(value):
    """Clean and normalize a text value."""
    if value is None:
        return ""
    s = str(value).strip()
    # Collapse multiple spaces
    s = re.sub(r"\s+", " ", s)
    return s


def get_year_int(value):
    """Convert a cell value to an integer year, or return None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    if s.isdigit() and len(s) == 4:
        return int(s)
    return None


# ──────────────────────────────────────────────────────────────────
# 1. All County Winners
# ──────────────────────────────────────────────────────────────────
def parse_all_county_winners():
    """
    Parse All County Winners.xlsx.

    Layout: Row 1 is headers. Data starts row 2.
    Columns:
      A: Years  B: Coach of the Year  C: MVP
      D: Offensive Player of Year  E: Defensive Player of the Year
      F: 1st Team Offense  G: 1st Team Defense
      H: 2nd Team Offense  I: 2nd Team Defense
      J: 3rd Team Offense  K: 3rd Team Defense

    Multi-row groups: A year appears in column A, then subsequent rows with
    column A empty belong to the same year. Each cell has one player name.
    """
    filepath = os.path.join(DOWNLOADS, "All County Winners.xlsx")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Column mapping: col_index -> (category, level, side)
    col_map = {
        2: ("Coach of the Year", "Coach of the Year", "N/A"),
        3: ("MVP", "MVP", "N/A"),
        4: ("Offensive Player of Year", "Player of the Year", "Offense"),
        5: ("Defensive Player of the Year", "Player of the Year", "Defense"),
        6: ("1st Team Offense", "1st Team", "Offense"),
        7: ("1st Team Defense", "1st Team", "Defense"),
        8: ("2nd Team Offense", "2nd Team", "Offense"),
        9: ("2nd Team Defense", "2nd Team", "Defense"),
        10: ("3rd Team Offense", "3rd Team", "Offense"),
        11: ("3rd Team Defense", "3rd Team", "Defense"),
    }

    honors = []
    current_year = None

    for r in range(2, ws.max_row + 1):
        year_val = get_year_int(ws.cell(row=r, column=1).value)
        if year_val is not None:
            current_year = year_val

        if current_year is None:
            continue

        for col_idx, (category, level, side) in col_map.items():
            cell_val = ws.cell(row=r, column=col_idx).value
            if is_empty(cell_val):
                continue

            # Some cells might have newlines (multiple names)
            text = str(cell_val).strip()
            names = text.split("\n")
            for name in names:
                name = name.strip()
                if name and name != "None":
                    honors.append({
                        "year": current_year,
                        "playerName": name,
                        "category": category,
                        "level": level,
                        "side": side,
                        "position": "",
                        "source": "All County Winners",
                    })

    wb.close()
    return honors


# ──────────────────────────────────────────────────────────────────
# 2. All League Winners
# ──────────────────────────────────────────────────────────────────
def parse_all_league_winners():
    """
    Parse All League Winners.xlsx.

    Columns:
      A: Years  B: MVP  C: Offensive Player of Year
      D: Defensive Player of the Year  E: Special Teams Player of the Year
      F: 1st Team Offense  G: 1st Team Defense
      H: 2nd Team Offense  I: 2nd Team Defense
      J: Honorable Mention Offense  K: Honorable Mention Defense

    Same multi-row grouping pattern as All County.
    """
    filepath = os.path.join(DOWNLOADS, "All League Winners.xlsx")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    col_map = {
        2: ("MVP", "MVP", "N/A"),
        3: ("Offensive Player of Year", "Player of the Year", "Offense"),
        4: ("Defensive Player of the Year", "Player of the Year", "Defense"),
        5: ("Special Teams Player of the Year", "Player of the Year", "Special Teams"),
        6: ("1st Team Offense", "1st Team", "Offense"),
        7: ("1st Team Defense", "1st Team", "Defense"),
        8: ("2nd Team Offense", "2nd Team", "Offense"),
        9: ("2nd Team Defense", "2nd Team", "Defense"),
        10: ("Honorable Mention Offense", "Honorable Mention", "Offense"),
        11: ("Honorable Mention Defense", "Honorable Mention", "Defense"),
    }

    honors = []
    current_year = None

    for r in range(2, ws.max_row + 1):
        year_val = get_year_int(ws.cell(row=r, column=1).value)
        if year_val is not None:
            current_year = year_val

        if current_year is None:
            continue

        for col_idx, (category, level, side) in col_map.items():
            cell_val = ws.cell(row=r, column=col_idx).value
            if is_empty(cell_val):
                continue

            text = str(cell_val).strip()
            names = text.split("\n")
            for name in names:
                name = name.strip()
                if name and name != "None":
                    honors.append({
                        "year": current_year,
                        "playerName": name,
                        "category": category,
                        "level": level,
                        "side": side,
                        "position": "",
                        "source": "All League Winners",
                    })

    wb.close()
    return honors


# ──────────────────────────────────────────────────────────────────
# 3. All State Recognition
# ──────────────────────────────────────────────────────────────────
def parse_all_state_recognition():
    """
    Parse All State Recognition.xlsx.

    Row 1: merged title row ("Mission Viejo All State Recognition")
    Row 2: headers -> Year, All CIF State Team, CIF SS, CIF Player of the Year
    Data: rows 3+

    Each cell contains a text description like:
      "All CIF 1st Team Offense  Nick Gilliam  Kicker Senior"
      "Division II First Team Defense Ryan Powdrell, LB Senior"
      "Offensive Player of the Year Robbie DuBois"

    Same multi-row year grouping. String "None" means no data.
    """
    filepath = os.path.join(DOWNLOADS, "All State Recognition.xlsx")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    col_categories = {
        2: "All CIF State Team",
        3: "CIF SS",
        4: "CIF Player of the Year",
    }

    honors = []
    current_year = None

    for r in range(3, ws.max_row + 1):
        year_val = get_year_int(ws.cell(row=r, column=1).value)
        if year_val is not None:
            current_year = year_val

        if current_year is None:
            continue

        for col_idx, category in col_categories.items():
            cell_val = ws.cell(row=r, column=col_idx).value
            if is_empty(cell_val):
                continue

            text = str(cell_val).strip()
            if text == "None":
                continue

            # The cell text is a full description; split on newlines if multiple
            entries = text.split("\n")
            for entry in entries:
                entry = entry.strip()
                if not entry or entry == "None":
                    continue

                # Try to extract side (Offense/Defense) from the text
                side = "N/A"
                if "Offense" in entry:
                    side = "Offense"
                elif "Defense" in entry:
                    side = "Defense"

                # Try to determine level from text
                level = category
                entry_lower = entry.lower()
                if "1st team" in entry_lower or "first team" in entry_lower:
                    level = "1st Team"
                elif "2nd team" in entry_lower or "second team" in entry_lower:
                    level = "2nd Team"
                elif "3rd team" in entry_lower or "third team" in entry_lower:
                    level = "3rd Team"
                elif "player of the year" in entry_lower:
                    level = "Player of the Year"
                elif "coach of the year" in entry_lower:
                    level = "Coach of the Year"
                elif "underclass" in entry_lower:
                    level = "Underclass"

                honors.append({
                    "year": current_year,
                    "playerName": entry,
                    "category": category,
                    "level": level,
                    "side": side,
                    "position": "",
                    "source": "All State Recognition",
                })

    wb.close()
    return honors


# ──────────────────────────────────────────────────────────────────
# 4. Cal-Hi All State Players
# ──────────────────────────────────────────────────────────────────
def parse_cal_hi_all_state():
    """
    Parse Cal Hi All State Players.xlsx.

    Row 1: headers -> Year, Cal-Hi All State Players
    Data: rows 2+. Each row has a year and a text description like:
      "All State 2nd Team Defense DL Matt Keneley"
      "All State Underclass Offense OL Drew Radovich"
    """
    filepath = os.path.join(DOWNLOADS, "Cal Hi All State Players.xlsx")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    honors = []

    for r in range(2, ws.max_row + 1):
        year_val = get_year_int(ws.cell(row=r, column=1).value)
        cell_val = ws.cell(row=r, column=2).value

        if is_empty(cell_val):
            continue

        text = str(cell_val).strip()
        if text == "None":
            continue

        year = year_val if year_val is not None else 0

        # Split on newlines if there are multiple entries
        entries = text.split("\n")
        for entry in entries:
            entry = entry.strip()
            if not entry or entry == "None":
                continue

            side = "N/A"
            if "Offense" in entry:
                side = "Offense"
            elif "Defense" in entry:
                side = "Defense"

            level = "Cal-Hi All State"
            entry_lower = entry.lower()
            if "first team" in entry_lower or "1st team" in entry_lower:
                level = "Cal-Hi 1st Team"
            elif "2nd team" in entry_lower or "second team" in entry_lower:
                level = "Cal-Hi 2nd Team"
            elif "3rd team" in entry_lower or "third team" in entry_lower:
                level = "Cal-Hi 3rd Team"
            elif "underclass" in entry_lower:
                level = "Cal-Hi Underclass"
            elif "sophomore" in entry_lower or "sophmore" in entry_lower:
                level = "Cal-Hi Sophomore"
            elif "position player" in entry_lower:
                level = "Cal-Hi Position Player of the Year"
            elif "player of the year" in entry_lower:
                level = "Cal-Hi Player of the Year"
            elif "preseason" in entry_lower:
                level = "Cal-Hi Preseason"
            elif "all decade" in entry_lower:
                level = "Cal-Hi All Decade Team"
            elif "medium school" in entry_lower or "small school" in entry_lower or "large school" in entry_lower:
                level = "Cal-Hi All State"

            honors.append({
                "year": year,
                "playerName": entry,
                "category": "Cal-Hi All State Players",
                "level": level,
                "side": side,
                "position": "",
                "source": "Cal Hi All State Players",
            })

    wb.close()
    return honors


# ──────────────────────────────────────────────────────────────────
# 5. CIF Cal Hi Records
# ──────────────────────────────────────────────────────────────────
def parse_cif_cal_hi_records():
    """
    Parse CIF Cal Hi Records.xlsx.

    Two columns: CIF Records, Cal Hi Records.
    These are structured as title/value pairs. A title row is followed
    by one or more value rows. We group them by detecting patterns:
    titles tend to not start with digits, values tend to start with digits
    or contain specific record data.
    """
    filepath = os.path.join(DOWNLOADS, "CIF Cal Hi Records.xlsx")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    def parse_column(col_idx, source_label):
        """Parse one column of records. Returns list of record dicts."""
        records = []
        current_title = None

        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if is_empty(val):
                continue

            text = clean_text(str(val))
            if not text:
                continue

            # Heuristic: a "title" row typically does NOT start with a digit
            # and does NOT contain a dash followed by a space and a name/team.
            # A "value" row typically starts with a digit or contains specific
            # record data patterns.
            is_value = bool(re.match(r"^\d", text))

            if is_value and current_title:
                records.append({
                    "recordTitle": current_title,
                    "recordValue": text,
                    "source": source_label,
                })
            else:
                # This is a new title
                current_title = text

        return records

    cif_records = parse_column(1, "CIF Records")
    cal_hi_records = parse_column(2, "Cal Hi Records")

    wb.close()
    return cif_records + cal_hi_records


# ──────────────────────────────────────────────────────────────────
# 6. Roster with sizes
# ──────────────────────────────────────────────────────────────────
def parse_roster():
    """
    Parse Copy of Roster with sizes.xlsx.

    Columns: Number, Last Name, First Name, Top Size, Bottom Size, Position
    Many rows at the end are empty (file has 998 rows but only ~70 with data).
    """
    filepath = os.path.join(DOWNLOADS, "Copy of Roster with sizes.xlsx")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    roster = []
    for r in range(2, ws.max_row + 1):
        last_name = ws.cell(row=r, column=2).value
        first_name = ws.cell(row=r, column=3).value

        # Skip rows without a name
        if is_empty(last_name) and is_empty(first_name):
            continue

        number = ws.cell(row=r, column=1).value
        if isinstance(number, (datetime.datetime, datetime.date)):
            # Excel sometimes misinterprets numbers as dates
            number = ""
        elif isinstance(number, float):
            number = int(number)

        top_size = ws.cell(row=r, column=4).value
        bottom_size = ws.cell(row=r, column=5).value
        position = ws.cell(row=r, column=6).value

        roster.append({
            "number": number if number is not None else "",
            "lastName": clean_text(last_name) if last_name else "",
            "firstName": clean_text(first_name) if first_name else "",
            "topSize": clean_text(top_size) if top_size else "",
            "bottomSize": clean_text(bottom_size) if bottom_size else "",
            "position": clean_text(position) if position else "",
        })

    wb.close()
    return roster


# ──────────────────────────────────────────────────────────────────
# 7. Contact Info CSV
# ──────────────────────────────────────────────────────────────────
def parse_contacts():
    """
    Parse Mission Viejo Football (2026 - 2027) Contact Info.csv.

    Columns: last, first, jersey #, email, cell,
             parent last, parent first, email, cell,
             emergency, last, first, cell

    Some parent fields contain multiple entries separated by newlines.
    """
    filepath = os.path.join(
        DOWNLOADS,
        "Mission Viejo Football (2026 - 2027) Contact Info.csv"
    )

    contacts = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader)  # skip header

        for row in reader:
            if not row or all(not cell.strip() for cell in row):
                continue

            # Pad row to 13 columns if short
            while len(row) < 13:
                row.append("")

            contact = {
                "lastName": row[0].strip(),
                "firstName": row[1].strip(),
                "jerseyNumber": row[2].strip(),
                "email": row[3].strip(),
                "cell": row[4].strip(),
                "parentLastName": row[5].strip(),
                "parentFirstName": row[6].strip(),
                "parentEmail": row[7].strip(),
                "parentCell": row[8].strip(),
                "emergencyRelation": row[9].strip(),
                "emergencyLastName": row[10].strip(),
                "emergencyFirstName": row[11].strip(),
                "emergencyCell": row[12].strip(),
            }
            contacts.append(contact)

    return contacts


# ──────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────
def main():
    print("Parsing All County Winners...")
    county_honors = parse_all_county_winners()
    print(f"  -> {len(county_honors)} honor records")

    print("Parsing All League Winners...")
    league_honors = parse_all_league_winners()
    print(f"  -> {len(league_honors)} honor records")

    print("Parsing All State Recognition...")
    state_honors = parse_all_state_recognition()
    print(f"  -> {len(state_honors)} honor records")

    print("Parsing Cal-Hi All State Players...")
    cal_hi_honors = parse_cal_hi_all_state()
    print(f"  -> {len(cal_hi_honors)} honor records")

    all_honors = county_honors + league_honors + state_honors + cal_hi_honors
    print(f"\nTotal honors: {len(all_honors)}")

    print("\nParsing CIF Cal Hi Records...")
    records = parse_cif_cal_hi_records()
    print(f"  -> {len(records)} record entries")

    print("\nParsing Roster...")
    roster = parse_roster()
    print(f"  -> {len(roster)} players")

    print("\nParsing Contact Info...")
    contacts = parse_contacts()
    print(f"  -> {len(contacts)} contacts")

    # Build output
    output = {
        "honors": all_honors,
        "records": records,
        "roster": roster,
        "contacts": contacts,
    }

    # Summary by source
    print("\n--- HONORS BREAKDOWN BY SOURCE ---")
    sources = {}
    for h in all_honors:
        src = h["source"]
        sources[src] = sources.get(src, 0) + 1
    for src, count in sorted(sources.items()):
        print(f"  {src}: {count}")

    # Summary by level
    print("\n--- HONORS BREAKDOWN BY LEVEL ---")
    levels = {}
    for h in all_honors:
        lvl = h["level"]
        levels[lvl] = levels.get(lvl, 0) + 1
    for lvl, count in sorted(levels.items()):
        print(f"  {lvl}: {count}")

    # Year range
    years = [h["year"] for h in all_honors if h["year"] > 0]
    if years:
        print(f"\nHonors year range: {min(years)} - {max(years)}")

    # Write output
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False, cls=SafeEncoder)

    print(f"\nOutput written to: {OUTPUT_FILE}")
    file_size = os.path.getsize(OUTPUT_FILE)
    print(f"File size: {file_size:,} bytes ({file_size / 1024:.1f} KB)")


if __name__ == "__main__":
    main()
